"""
src/report_builder.py  —  Build the formatted Excel MIS report
Generates a professional multi-sheet Excel workbook with:
  - Executive Summary (AI commentary + KPI scorecards)
  - P&L Summary
  - Department Breakdown
  - Anomaly Log
  - Rolling Trends (raw data for Power BI)
"""

import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import config


# ── Color palette ─────────────────────────────────────────────────────────────
C_DARK_GREEN  = "FF085041"
C_MID_GREEN   = "FF1D9E75"
C_LIGHT_GREEN = "FFE1F5EE"
C_AMBER       = "FFFAEEDA"
C_AMBER_DARK  = "FF633806"
C_RED_LIGHT   = "FFFCEBEB"
C_RED_DARK    = "FFA32D2D"
C_GRAY_LIGHT  = "FFF1EFE8"
C_GRAY_MID    = "FFD3D1C7"
C_WHITE       = "FFFFFFFF"
C_TEXT_DARK   = "FF1A1A1A"
C_TEXT_MID    = "FF444441"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="FF000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def _border(style="thin") -> Border:
    s = Side(style=style, color="FFD3D1C7")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width

def _header_row(ws, row: int, values: list, col_start: int = 1):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=val)
        cell.fill    = _fill(C_DARK_GREEN)
        cell.font    = _font(bold=True, size=10, color="FFFFFFFF")
        cell.alignment = _align("center")
        cell.border  = _border()

def _variance_cell(ws, row: int, col: int, value: float, is_pct: bool = True):
    cell = ws.cell(row=row, column=col, value=round(value, 2))
    if value < 0:
        cell.fill = _fill(C_RED_LIGHT)
        cell.font = _font(bold=True, color=C_RED_DARK)
    elif value > 0:
        cell.fill = _fill(C_LIGHT_GREEN)
        cell.font = _font(bold=True, color=C_DARK_GREEN)
    else:
        cell.font = _font()
    cell.number_format = "0.0%" if is_pct else "#,##0.00"
    cell.alignment = _align("center")
    cell.border    = _border()
    return cell


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_executive_summary(ws, kpis: dict, commentary: str, risk_flags: list):
    ws.sheet_view.showGridLines = False

    # Title bar
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value     = f"{config.COMPANY_NAME}  —  Weekly MIS Report"
    title.font      = _font(bold=True, size=16, color="FFFFFFFF")
    title.fill      = _fill(C_DARK_GREEN)
    title.alignment = _align("center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value     = f"Period: {kpis.get('period_start', 'N/A')} to {kpis.get('period_end', 'N/A')}   |   Generated: {datetime.today().strftime('%d %b %Y %H:%M')}"
    sub.font      = _font(size=10, color=C_TEXT_MID, italic=True)
    sub.fill      = _fill(C_LIGHT_GREEN)
    sub.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    # KPI Scorecards (row 4–8)
    scorecard_data = [
        ("Total Revenue",      f"£{kpis.get('total_revenue', 0):,.0f}",      None),
        ("Gross Margin",       f"{kpis.get('gross_margin_pct', 0):.1f}%",     None),
        ("Net Profit",         f"£{kpis.get('net_profit', 0):,.0f}",          None),
        ("Budget Variance",    f"{kpis.get('budget_variance_pct', 0):+.1f}%", kpis.get('budget_variance_pct', 0)),
        ("MoM Revenue Δ",      f"{kpis.get('mom_revenue_change_pct', 0):+.1f}%", kpis.get('mom_revenue_change_pct', 0)),
        ("SLA Breach Rate",    f"{kpis.get('sla_breach_rate_pct', 0):.1f}%",  None),
        ("Total Budget",       f"£{kpis.get('total_budget', 0):,.0f}",        None),
        ("Rows Processed",     f"{kpis.get('rows_processed', 0):,}",          None),
    ]

    ws.row_dimensions[3].height = 10
    for col_idx, (label, value, raw) in enumerate(scorecard_data, start=1):
        col = get_column_letter(col_idx)
        # Label row
        lc = ws[f"{col}4"]
        lc.value     = label
        lc.font      = _font(size=9, color=C_TEXT_MID)
        lc.fill      = _fill(C_GRAY_LIGHT)
        lc.alignment = _align("center")
        lc.border    = _border()
        ws.row_dimensions[4].height = 16

        # Value row
        vc = ws[f"{col}5"]
        vc.value     = value
        vc.font      = _font(bold=True, size=13)
        vc.alignment = _align("center")
        vc.border    = _border()
        ws.row_dimensions[5].height = 26

        if raw is not None:
            vc.fill = _fill(C_RED_LIGHT if raw < 0 else C_LIGHT_GREEN)
            vc.font = _font(bold=True, size=13,
                            color=C_RED_DARK if raw < 0 else C_DARK_GREEN)

    ws.row_dimensions[6].height = 10

    # AI Commentary
    ws.merge_cells("A7:H7")
    hdr = ws["A7"]
    hdr.value     = "AI Executive Commentary"
    hdr.font      = _font(bold=True, size=11, color="FFFFFFFF")
    hdr.fill      = _fill(C_MID_GREEN)
    hdr.alignment = _align("left")
    ws.row_dimensions[7].height = 20

    ws.merge_cells("A8:H14")
    cc = ws["A8"]
    cc.value     = commentary
    cc.font      = _font(size=10)
    cc.alignment = _align("left", wrap=True)
    cc.fill      = _fill(C_AMBER)
    cc.border    = _border()
    for r in range(8, 15):
        ws.row_dimensions[r].height = 22

    # Risk flags
    ws.row_dimensions[15].height = 8
    ws.merge_cells("A16:H16")
    rh = ws["A16"]
    rh.value     = "Risk Flags"
    rh.font      = _font(bold=True, size=11, color="FFFFFFFF")
    rh.fill      = _fill(C_RED_DARK[2:] and "FFA32D2D")
    rh.alignment = _align("left")
    ws.row_dimensions[16].height = 20

    if risk_flags:
        for i, flag in enumerate(risk_flags, start=17):
            ws.merge_cells(f"A{i}:H{i}")
            fc = ws[f"A{i}"]
            fc.value     = f"[{flag['severity'].upper()}]  {flag['flag']}  —  {flag['detail']}"
            fc.font      = _font(size=10, bold=(flag["severity"] == "high"))
            fc.fill      = _fill(C_RED_LIGHT)
            fc.alignment = _align("left")
            fc.border    = _border()
            ws.row_dimensions[i].height = 18
    else:
        ws.merge_cells("A17:H17")
        nc = ws["A17"]
        nc.value     = "No critical risk flags this period."
        nc.font      = _font(size=10, color=C_DARK_GREEN)
        nc.fill      = _fill(C_LIGHT_GREEN)
        nc.alignment = _align("left")

    for col in range(1, 9):
        _set_col_width(ws, get_column_letter(col), 18)


def _build_pl_sheet(ws, kpis: dict):
    ws.sheet_view.showGridLines = False
    headers = ["P&L Line Item", "This Week (£)", "Budget (£)", "Variance (£)", "Margin %"]
    _header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 22

    pl_rows = [
        ("Revenue",           kpis.get("total_revenue", 0),    kpis.get("total_budget", 0)),
        ("Cost of Goods Sold",kpis.get("cogs", 0),             None),
        ("Gross Profit",      kpis.get("gross_profit", 0),     None),
        ("Operating Expenses",kpis.get("opex", 0),             None),
        ("Net Profit",        kpis.get("net_profit", 0),        None),
    ]

    for r, (label, value, budget) in enumerate(pl_rows, start=2):
        bold = label in ("Gross Profit", "Net Profit", "Revenue")
        ws.cell(r, 1, label).font = _font(bold=bold, size=10)
        ws.cell(r, 1).border      = _border()

        val_cell = ws.cell(r, 2, round(value, 2))
        val_cell.number_format = "#,##0.00"
        val_cell.font   = _font(bold=bold, size=10)
        val_cell.border = _border()

        if budget is not None:
            bud_cell = ws.cell(r, 3, round(budget, 2))
            bud_cell.number_format = "#,##0.00"
            bud_cell.border = _border()
            variance = value - budget
            _variance_cell(ws, r, 4, variance, is_pct=False)
        else:
            ws.cell(r, 3, "—").border = _border()
            ws.cell(r, 4, "—").border = _border()

        if label in ("Gross Profit",):
            rev = kpis.get("total_revenue", 1)
            margin_cell = ws.cell(r, 5, round(value / rev * 100 if rev else 0, 2))
            margin_cell.number_format = "0.0%"
            margin_cell.border = _border()
        else:
            ws.cell(r, 5, "—").border = _border()

        ws.row_dimensions[r].height = 20

    for col, width in zip("ABCDE", [28, 18, 18, 18, 14]):
        _set_col_width(ws, col, width)


def _build_dept_sheet(ws, dept_df: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    if dept_df is None or dept_df.empty:
        ws["A1"] = "No department data available."
        return

    headers = ["Department", "Revenue (£)", "Budget (£)", "Actuals (£)",
               "Variance %", "Gross Margin %", "SLA Breach %"]
    _header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(dept_df.itertuples(), start=2):
        ws.cell(r, 1, row.department).border  = _border()
        ws.cell(r, 2, round(row.revenue, 0)).number_format = "#,##0"
        ws.cell(r, 2).border = _border()
        ws.cell(r, 3, round(row.budget, 0)).number_format = "#,##0"
        ws.cell(r, 3).border = _border()
        ws.cell(r, 4, round(row.actuals, 0)).number_format = "#,##0"
        ws.cell(r, 4).border = _border()
        _variance_cell(ws, r, 5, row.variance_pct)
        _variance_cell(ws, r, 6, row.gross_margin_pct, is_pct=False)
        _variance_cell(ws, r, 7, row.sla_breach_rate_pct, is_pct=False)
        ws.row_dimensions[r].height = 18

    for col, width in zip("ABCDEFG", [20, 16, 16, 16, 14, 16, 14]):
        _set_col_width(ws, col, width)


def _build_anomaly_sheet(ws, anomalies: list, anomaly_summary: str):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value     = "AI Anomaly Analysis"
    title.font      = _font(bold=True, size=11, color="FFFFFFFF")
    title.fill      = _fill(C_DARK_GREEN)
    title.alignment = _align("left")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("A2:G4")
    sc = ws["A2"]
    sc.value     = anomaly_summary
    sc.font      = _font(size=10, italic=True)
    sc.alignment = _align("left", wrap=True)
    sc.fill      = _fill(C_AMBER)
    sc.border    = _border()
    for r in range(2, 5):
        ws.row_dimensions[r].height = 18

    if anomalies:
        headers = ["Date", "Department", "Metric", "Value", "Z-Score", "Direction", "Severity"]
        _header_row(ws, 5, headers)

        for r, a in enumerate(anomalies, start=6):
            sev_fill = C_RED_LIGHT if a["severity"] == "high" else C_AMBER
            for col, val in enumerate([
                a["date"], a["department"], a["metric"],
                a["value"], a["z_score"], a["direction"], a["severity"].upper()
            ], start=1):
                cell = ws.cell(r, col, val)
                cell.fill      = _fill(sev_fill)
                cell.border    = _border()
                cell.alignment = _align("center")
                cell.font      = _font(size=10)
            ws.row_dimensions[r].height = 18

    for col, width in zip("ABCDEFG", [14, 18, 22, 14, 10, 12, 10]):
        _set_col_width(ws, col, width)


def _build_trends_sheet(ws, trends_df: pd.DataFrame):
    """Raw trend data sheet — consumed by Power BI directly."""
    ws.sheet_view.showGridLines = False
    if trends_df is None or trends_df.empty:
        ws["A1"] = "No trend data available."
        return

    trends_df = trends_df.copy()
    trends_df["week"] = trends_df["week"].astype(str)

    headers = list(trends_df.columns)
    _header_row(ws, 1, headers)

    for r, row in enumerate(trends_df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            cell.border    = _border()
            cell.alignment = _align("center")
            cell.font      = _font(size=10)
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
        ws.row_dimensions[r].height = 18

    for i, col in enumerate(headers):
        _set_col_width(ws, get_column_letter(i + 1), max(len(col) + 4, 14))


# ── Master report builder ─────────────────────────────────────────────────────

def build_excel_report(
    kpis:             dict,
    commentary:       str,
    anomaly_summary:  str,
    risk_flags:       list,
    anomalies:        list,
    dept_df,
    trends_df,
    output_path:      str = None,
) -> str:
    """
    Build the complete multi-sheet Excel MIS report.
    Returns the path of the saved file.
    """
    if output_path is None:
        date_str    = datetime.today().strftime("%Y-%m-%d")
        filename    = f"MIS_Report_{date_str}.xlsx"
        output_path = os.path.join(config.REPORT_OUTPUT_DIR, filename)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()

    # 1. Executive Summary
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    _build_executive_summary(ws_exec, kpis, commentary, risk_flags)

    # 2. P&L Summary
    ws_pl = wb.create_sheet("P&L Summary")
    _build_pl_sheet(ws_pl, kpis)

    # 3. Department Breakdown
    ws_dept = wb.create_sheet("Departments")
    _build_dept_sheet(ws_dept, dept_df)

    # 4. Anomaly Log
    ws_anomaly = wb.create_sheet("Anomalies")
    _build_anomaly_sheet(ws_anomaly, anomalies, anomaly_summary)

    # 5. Rolling Trends (Power BI source)
    ws_trends = wb.create_sheet("Trends (Power BI)")
    _build_trends_sheet(ws_trends, trends_df)

    wb.save(output_path)
    print(f"  Report saved: {output_path}")
    return output_path
